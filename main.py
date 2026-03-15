import datetime
import os
from fastapi import FastAPI, Depends, HTTPException, Request
from fastapi.staticfiles import StaticFiles
from fastapi.responses import FileResponse, JSONResponse, StreamingResponse
from fastapi.middleware.cors import CORSMiddleware
from sqlalchemy.orm import Session
from typing import List
from pydantic import BaseModel

import models
import schemas
import auth
from database import engine, get_db


# Create DB Tables
models.Base.metadata.create_all(bind=engine)

app = FastAPI(title="Narayana Dairy Farm API")

# CORS — allows both partners to connect from different devices on LAN
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# Static files
if not os.path.exists("static"):
    os.makedirs("static")
app.mount("/static", StaticFiles(directory="static"), name="static")


# ─── No-Cache Middleware for JS/HTML/CSS ─────────────────────────────────────
@app.middleware("http")
async def no_cache_static_middleware(request: Request, call_next):
    response = await call_next(request)
    path = request.url.path
    # Force browser to always revalidate JS, HTML, CSS files
    if path.endswith(('.js', '.html', '.css')) or path == '/':
        response.headers['Cache-Control'] = 'no-cache, no-store, must-revalidate'
        response.headers['Pragma'] = 'no-cache'
        response.headers['Expires'] = '0'
    return response


# ─── Auth Middleware ──────────────────────────────────────────────────────────
EXEMPT_PREFIXES = ["/static/", "/api/auth/", "/docs", "/openapi.json", "/redoc", "/favicon.ico"]

@app.middleware("http")
async def auth_middleware(request: Request, call_next):
    path = request.url.path
    if path == "/" or any(path.startswith(e) for e in EXEMPT_PREFIXES):
        return await call_next(request)
    header = request.headers.get("Authorization", "")
    if not header.startswith("Bearer "):
        return JSONResponse(status_code=401, content={"detail": "Not authenticated"})
    payload = auth.decode_access_token(header[7:])
    if not payload:
        return JSONResponse(status_code=401, content={"detail": "Session expired. Please login again."})
    request.state.user = payload
    return await call_next(request)


# ─── Pages ───────────────────────────────────────────────────────────────────
@app.get("/", include_in_schema=False)
def serve_root():
    return FileResponse("static/index.html")


# ─── Auth Endpoints (Public) ──────────────────────────────────────────────────
class OTPRequest(BaseModel):
    name: str

class OTPVerify(BaseModel):
    name: str
    otp: str

@app.get("/api/auth/partners")
def get_partners():
    """Public endpoint — returns partner names only (not emails) for the login page."""
    config = auth.load_config()
    names = [p["name"] for p in config.get("partners", [])]
    return {"partners": names}

@app.post("/api/auth/send-otp")
def send_otp(body: OTPRequest):
    config  = auth.load_config()
    # Look up partner by name
    partner = next((p for p in config.get("partners", []) if p["name"] == body.name), None)
    if not partner:
        raise HTTPException(status_code=404, detail="Partner not found.")

    otp = auth.generate_otp()
    # Key OTP store by name (not email) since email is internal
    auth.save_otp(body.name, otp)

    smtp_cfg      = config.get("smtp", {})
    smtp_ready    = bool(smtp_cfg.get("username") and smtp_cfg.get("password"))
    delivery_email = config.get("otp_delivery_email") or partner.get("email", "")

    if smtp_ready and delivery_email:
        try:
            auth.send_otp_email(partner, otp, smtp_cfg, delivery_email=delivery_email)
            masked = delivery_email[:4] + "****" + delivery_email[delivery_email.index("@"):]
            return {"status": "sent", "message": f"OTP sent to {masked}"}
        except Exception as ex:
            raise HTTPException(status_code=500,
                detail=f"Email delivery failed: {str(ex)}")
    else:
        return {"status": "dev", "dev_otp": otp,
                "message": "SMTP not configured — OTP shown on screen (dev mode)."}


@app.post("/api/auth/verify-otp")
def verify_otp(body: OTPVerify):
    config  = auth.load_config()
    partner = next((p for p in config.get("partners", []) if p["name"] == body.name), None)
    if not partner:
        raise HTTPException(status_code=404, detail="Partner not found.")
    # OTP is stored by name
    ok, msg = auth.verify_otp_code(body.name, body.otp)
    if not ok:
        raise HTTPException(status_code=400, detail=msg)
    token = auth.create_access_token(partner.get("email", body.name), partner["name"])

    # ── Record login ────────────────────────────────────────────────────
    now = datetime.datetime.now()
    log = models.LoginLog(
        partner_name=partner["name"],
        login_date=now.date(),
        login_time=now.strftime("%H:%M:%S"),
        ip_address=None   # IP not available without Request injection here
    )
    # We need a db session - use a fresh one
    from database import SessionLocal
    _db = SessionLocal()
    try:
        _db.add(log); _db.commit()
    finally:
        _db.close()

    return {"access_token": token, "token_type": "bearer",
            "partner_name": partner["name"]}



# ─── Buffalo Endpoints ────────────────────────────────────────────────────────
@app.get("/api/buffaloes", response_model=List[schemas.Buffalo])
def get_buffaloes(db: Session = Depends(get_db)):
    return db.query(models.Buffalo).all()

@app.post("/api/buffaloes", response_model=schemas.Buffalo)
def create_buffalo(buffalo: schemas.BuffaloCreate, db: Session = Depends(get_db)):
    db_buffalo = models.Buffalo(**buffalo.dict())
    db.add(db_buffalo); db.commit(); db.refresh(db_buffalo)
    return db_buffalo

@app.delete("/api/buffaloes/{buffalo_id}")
def delete_buffalo(buffalo_id: str, db: Session = Depends(get_db)):
    b = db.query(models.Buffalo).filter(models.Buffalo.id == buffalo_id).first()
    if not b: raise HTTPException(404, "Buffalo not found")
    db.delete(b); db.commit()
    return {"message": "Buffalo deleted"}


# ─── Milk Production ──────────────────────────────────────────────────────────
@app.get("/api/milk/daily", response_model=List[schemas.MilkProduction])
def get_daily_milk(date: datetime.date = None, db: Session = Depends(get_db)):
    target = date or datetime.date.today()
    return db.query(models.MilkProduction).filter(models.MilkProduction.date == target).all()

@app.get("/api/milk/all", response_model=List[schemas.MilkProduction])
def get_all_milk(db: Session = Depends(get_db)):
    return db.query(models.MilkProduction).order_by(models.MilkProduction.date.desc()).all()

@app.post("/api/milk", response_model=schemas.MilkProduction)
def create_milk_record(milk: schemas.MilkProductionCreate, db: Session = Depends(get_db)):
    db_milk = models.MilkProduction(**milk.dict())
    db_milk.total_milk_liters = db_milk.morning_milk_liters + db_milk.evening_milk_liters
    db.add(db_milk); db.commit(); db.refresh(db_milk)
    return db_milk

@app.delete("/api/milk/{milk_id}")
def delete_milk_record(milk_id: str, db: Session = Depends(get_db)):
    m = db.query(models.MilkProduction).filter(models.MilkProduction.id == milk_id).first()
    if not m: raise HTTPException(404, "Milk record not found")
    db.delete(m); db.commit()
    return {"message": "Milk record deleted"}


# ─── Milk Sales (Income) ──────────────────────────────────────────────────────
@app.post("/api/sales", response_model=schemas.MilkSales)
def create_sales_record(sales: schemas.MilkSalesCreate, db: Session = Depends(get_db)):
    db_sales = models.MilkSales(**sales.dict())
    db_sales.total_income = db_sales.quantity_supplied_liters * db_sales.price_per_liter
    db.add(db_sales); db.commit(); db.refresh(db_sales)
    return db_sales

@app.get("/api/sales", response_model=List[schemas.MilkSales])
def get_sales_records(db: Session = Depends(get_db)):
    return db.query(models.MilkSales).order_by(models.MilkSales.date.desc()).all()

@app.delete("/api/sales/{sale_id}")
def delete_sales_record(sale_id: str, db: Session = Depends(get_db)):
    s = db.query(models.MilkSales).filter(models.MilkSales.id == sale_id).first()
    if not s: raise HTTPException(404, "Sale record not found")
    db.delete(s); db.commit()
    return {"message": "Sale record deleted"}


# ─── Expenses ─────────────────────────────────────────────────────────────────
@app.post("/api/expenses", response_model=schemas.Expense)
def create_expense(expense: schemas.ExpenseCreate, db: Session = Depends(get_db)):
    db_expense = models.Expense(**expense.dict())
    db.add(db_expense); db.commit(); db.refresh(db_expense)
    return db_expense

@app.get("/api/expenses", response_model=List[schemas.Expense])
def get_expenses(db: Session = Depends(get_db)):
    return db.query(models.Expense).all()

@app.delete("/api/expenses/{expense_id}")
def delete_expense(expense_id: str, db: Session = Depends(get_db)):
    e = db.query(models.Expense).filter(models.Expense.id == expense_id).first()
    if not e: raise HTTPException(404, "Expense not found")
    db.delete(e); db.commit()
    return {"message": "Expense deleted"}


# ─── Health Records ───────────────────────────────────────────────────────────
@app.post("/api/health", response_model=schemas.HealthRecord)
def create_health_record(health: schemas.HealthRecordCreate, db: Session = Depends(get_db)):
    db_health = models.HealthRecord(**health.dict())
    db.add(db_health); db.commit(); db.refresh(db_health)
    return db_health

@app.get("/api/health", response_model=List[schemas.HealthRecord])
def get_health_records(db: Session = Depends(get_db)):
    return db.query(models.HealthRecord).order_by(models.HealthRecord.date.desc()).all()

@app.delete("/api/health/{record_id}")
def delete_health_record(record_id: str, db: Session = Depends(get_db)):
    h = db.query(models.HealthRecord).filter(models.HealthRecord.id == record_id).first()
    if not h: raise HTTPException(404, "Health record not found")
    db.delete(h); db.commit()
    return {"message": "Health record deleted"}

@app.get("/api/health/alerts", response_model=List[schemas.HealthRecord])
def get_health_alerts(db: Session = Depends(get_db)):
    today    = datetime.date.today()
    in_7days = today + datetime.timedelta(days=7)
    return db.query(models.HealthRecord).filter(
        models.HealthRecord.next_due_date != None,
        models.HealthRecord.next_due_date <= in_7days
    ).all()


# ─── Dashboard ────────────────────────────────────────────────────────────────
@app.get("/api/dashboard", response_model=schemas.DashboardSummary)
def get_dashboard_summary(db: Session = Depends(get_db)):
    from sqlalchemy import text
    today          = datetime.date.today()
    first_of_month = today.replace(day=1)
    in_7days       = today + datetime.timedelta(days=7)

    total_buffaloes    = db.execute(text("SELECT COUNT(*) FROM buffaloes")).scalar() or 0
    pregnant_buffaloes = db.execute(text("SELECT COUNT(*) FROM buffaloes WHERE pregnancy_status=1")).scalar() or 0
    total_vendors      = db.execute(text("SELECT COUNT(*) FROM vendors")).scalar() or 0
    upcoming_vacc      = db.execute(text(
        "SELECT COUNT(*) FROM health_records WHERE record_type='VACCINATION' AND next_due_date IS NOT NULL AND next_due_date<=:d"
    ), {"d": str(in_7days)}).scalar() or 0
    t_milk   = sum(r[0] for r in db.execute(text("SELECT total_milk_liters FROM milk_production WHERE date=:d"), {"d": str(today)}).fetchall() if r[0])
    t_exp    = sum(r[0] for r in db.execute(text("SELECT amount FROM expenses WHERE date>=:d"), {"d": str(first_of_month)}).fetchall() if r[0])
    t_sales  = sum(r[0] for r in db.execute(text("SELECT total_income FROM milk_sales WHERE date>=:d"), {"d": str(first_of_month)}).fetchall() if r[0])
    return schemas.DashboardSummary(
        total_buffaloes=total_buffaloes, pregnant_buffaloes=pregnant_buffaloes,
        total_milk_today=t_milk, upcoming_vaccinations=upcoming_vacc,
        profit_loss_current_month=(t_sales - t_exp), total_vendors=total_vendors
    )


# ─── Vendors ──────────────────────────────────────────────────────────────────
@app.get("/api/vendors", response_model=List[schemas.Vendor])
def get_vendors(db: Session = Depends(get_db)):
    return db.query(models.Vendor).order_by(models.Vendor.name).all()

@app.post("/api/vendors", response_model=schemas.Vendor)
def create_vendor(vendor: schemas.VendorCreate, db: Session = Depends(get_db)):
    data = vendor.dict()
    if not data.get("transaction_date"):
        data["transaction_date"] = datetime.date.today()
    db_vendor = models.Vendor(**data)
    db.add(db_vendor); db.commit(); db.refresh(db_vendor)
    return db_vendor

@app.get("/api/vendors/{vendor_id}", response_model=schemas.Vendor)
def get_vendor(vendor_id: str, db: Session = Depends(get_db)):
    v = db.query(models.Vendor).filter(models.Vendor.id == vendor_id).first()
    if not v: raise HTTPException(404, "Vendor not found")
    return v

@app.delete("/api/vendors/{vendor_id}")
def delete_vendor(vendor_id: str, db: Session = Depends(get_db)):
    v = db.query(models.Vendor).filter(models.Vendor.id == vendor_id).first()
    if not v: raise HTTPException(404, "Vendor not found")
    # Explicitly remove linked payments first (SQLite doesn't auto-cascade)
    db.query(models.VendorPayment).filter(models.VendorPayment.vendor_id == vendor_id).delete()
    db.delete(v)
    db.commit()
    return {"message": "Vendor deleted"}


# ─── Finance Ledger Excel Download ───────────────────────────────────────────
@app.get("/api/finance/download-ledger")
def download_ledger(db: Session = Depends(get_db)):
    import io
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    sales    = db.query(models.MilkSales).order_by(models.MilkSales.date).all()
    expenses = db.query(models.Expense).order_by(models.Expense.date).all()

    # Build unified entries sorted by date
    entries = []
    for s in sales:
        entries.append({
            "date":    str(s.date),
            "type":    "INCOME",
            "desc":    f"Milk Sale - {s.milk_center_name}",
            "details": f"{s.quantity_supplied_liters} L @ Rs.{s.price_per_liter}/L",
            "income":  float(s.total_income or 0),
            "expense": 0.0,
        })
    for e in expenses:
        entries.append({
            "date":    str(e.date),
            "type":    "EXPENSE",
            "desc":    e.category,
            "details": e.description or "",
            "income":  0.0,
            "expense": float(e.amount or 0),
        })
    entries.sort(key=lambda x: x["date"])

    # Create workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Finance Ledger"

    # Header row
    headers = ["Date", "Type", "Description", "Details", "Income (Rs.)", "Expense (Rs.)", "Balance (Rs.)"]
    ws.append(headers)
    hdr_fill   = PatternFill("solid", fgColor="064E3B")
    hdr_font   = Font(bold=True, color="FFFFFF", size=11)
    hdr_align  = Alignment(horizontal="center", vertical="center")
    thin_side  = Side(style="thin", color="CCCCCC")
    thin_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
    for cell in ws[1]:
        cell.fill   = hdr_fill
        cell.font   = hdr_font
        cell.alignment = hdr_align
        cell.border = thin_border
    ws.row_dimensions[1].height = 22

    # Data rows
    balance = 0.0
    inc_fill = PatternFill("solid", fgColor="D1FAE5")  # light green
    exp_fill = PatternFill("solid", fgColor="FEE2E2")  # light red
    for entry in entries:
        balance += entry["income"] - entry["expense"]
        row = [
            entry["date"],
            entry["type"],
            entry["desc"],
            entry["details"],
            round(entry["income"], 2)  if entry["income"]  > 0 else "",
            round(entry["expense"], 2) if entry["expense"] > 0 else "",
            round(balance, 2),
        ]
        ws.append(row)
        fill = inc_fill if entry["type"] == "INCOME" else exp_fill
        for cell in ws[ws.max_row]:
            cell.fill   = fill
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center")

    # Auto-width columns
    col_widths = [12, 10, 32, 30, 14, 14, 14]
    for i, col in enumerate(ws.columns):
        ws.column_dimensions[col[0].column_letter].width = col_widths[i]

    # Freeze header row
    ws.freeze_panes = "A2"

    # Stream as .xlsx
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    today = datetime.date.today().isoformat()
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="Finance_Ledger_{today}.xlsx"'}
    )


# ─── Helper: make workbook style elements ─────────────────────────────────────
def _excel_header_style():
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    hdr_fill  = PatternFill("solid", fgColor="064E3B")
    hdr_font  = Font(bold=True, color="FFFFFF", size=11)
    hdr_align = Alignment(horizontal="center", vertical="center")
    thin      = Side(style="thin", color="CCCCCC")
    border    = Border(left=thin, right=thin, top=thin, bottom=thin)
    return hdr_fill, hdr_font, hdr_align, border

def _apply_row_style(ws, fill, border):
    from openpyxl.styles import Alignment
    for cell in ws[ws.max_row]:
        cell.fill   = fill
        cell.border = border
        cell.alignment = Alignment(vertical="center")


# ─── Download: Buffalo List ───────────────────────────────────────────────────
@app.get("/api/download/buffaloes")
def download_buffaloes(db: Session = Depends(get_db)):
    import io, openpyxl
    from openpyxl.styles import PatternFill

    buffaloes = db.query(models.Buffalo).order_by(models.Buffalo.tag_number).all()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Buffalo List"

    headers = ["Tag Number", "Name", "Breed", "Date of Birth", "Lactation No.",
               "Pregnant?", "Notes", "Added On"]
    ws.append(headers)
    hdr_fill, hdr_font, hdr_align, border = _excel_header_style()
    for cell in ws[1]:
        cell.fill = hdr_fill; cell.font = hdr_font
        cell.alignment = hdr_align; cell.border = border
    ws.row_dimensions[1].height = 22

    alt1 = PatternFill("solid", fgColor="F0FDF4")
    alt2 = PatternFill("solid", fgColor="FFFFFF")
    for i, b in enumerate(buffaloes):
        row = [
            b.tag_number, b.name or "", b.breed,
            str(b.date_of_birth) if b.date_of_birth else "",
            b.lactation_number,
            "Yes" if b.pregnancy_status else "No",
            b.notes or "",
            str(b.created_at.date()) if b.created_at else ""
        ]
        ws.append(row)
        _apply_row_style(ws, alt1 if i % 2 == 0 else alt2, border)

    col_widths = [14, 16, 16, 14, 14, 12, 30, 14]
    for i, col in enumerate(ws.columns):
        ws.column_dimensions[col[0].column_letter].width = col_widths[i]
    ws.freeze_panes = "A2"

    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    today = datetime.date.today().isoformat()
    return StreamingResponse(buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="Buffalo_List_{today}.xlsx"'})


# ─── Download: Milk Production ────────────────────────────────────────────────
@app.get("/api/download/milk")
def download_milk(db: Session = Depends(get_db)):
    import io, openpyxl
    from openpyxl.styles import PatternFill

    records = db.query(models.MilkProduction).order_by(models.MilkProduction.date.desc()).all()
    # Build buffalo tag map
    bufmap = {b.id: b.tag_number for b in db.query(models.Buffalo).all()}

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Milk Production"

    headers = ["Date", "Buffalo Tag", "Morning (L)", "Evening (L)", "Total (L)", "Recorded On"]
    ws.append(headers)
    hdr_fill, hdr_font, hdr_align, border = _excel_header_style()
    for cell in ws[1]:
        cell.fill = hdr_fill; cell.font = hdr_font
        cell.alignment = hdr_align; cell.border = border
    ws.row_dimensions[1].height = 22

    alt1 = PatternFill("solid", fgColor="ECFDF5")
    alt2 = PatternFill("solid", fgColor="FFFFFF")
    for i, r in enumerate(records):
        row = [
            str(r.date), bufmap.get(r.buffalo_id, r.buffalo_id[:8]),
            round(r.morning_milk_liters, 2),
            round(r.evening_milk_liters, 2),
            round(r.total_milk_liters, 2),
            str(r.created_at.date()) if r.created_at else ""
        ]
        ws.append(row)
        _apply_row_style(ws, alt1 if i % 2 == 0 else alt2, border)

    col_widths = [12, 16, 14, 14, 12, 14]
    for i, col in enumerate(ws.columns):
        ws.column_dimensions[col[0].column_letter].width = col_widths[i]
    ws.freeze_panes = "A2"

    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    today = datetime.date.today().isoformat()
    return StreamingResponse(buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="Milk_Production_{today}.xlsx"'})


# ─── Download: Health Records ─────────────────────────────────────────────────
@app.get("/api/download/health")
def download_health(db: Session = Depends(get_db)):
    import io, openpyxl
    from openpyxl.styles import PatternFill

    records = db.query(models.HealthRecord).order_by(models.HealthRecord.date.desc()).all()
    bufmap  = {b.id: b.tag_number for b in db.query(models.Buffalo).all()}

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Health Records"

    headers = ["Date", "Buffalo Tag", "Record Type", "Details", "Next Due Date", "Recorded On"]
    ws.append(headers)
    hdr_fill, hdr_font, hdr_align, border = _excel_header_style()
    for cell in ws[1]:
        cell.fill = hdr_fill; cell.font = hdr_font
        cell.alignment = hdr_align; cell.border = border
    ws.row_dimensions[1].height = 22

    vacc_fill  = PatternFill("solid", fgColor="EEF2FF")   # light indigo — vaccinations
    treat_fill = PatternFill("solid", fgColor="FFF7ED")   # light orange — treatment
    gen_fill   = PatternFill("solid", fgColor="F0FDF4")   # light green  — general

    type_fills = {"VACCINATION": vacc_fill, "PREGNANCY_CHECK": PatternFill("solid", fgColor="FDF2F8"), "CHECKUP": gen_fill}

    for r in records:
        row = [
            str(r.date),
            bufmap.get(r.buffalo_id, r.buffalo_id[:8]),
            r.record_type,
            r.details or "",
            str(r.next_due_date) if r.next_due_date else "",
            str(r.created_at.date()) if r.created_at else ""
        ]
        ws.append(row)
        fill = type_fills.get(r.record_type, treat_fill)
        _apply_row_style(ws, fill, border)

    col_widths = [12, 16, 20, 40, 14, 14]
    for i, col in enumerate(ws.columns):
        ws.column_dimensions[col[0].column_letter].width = col_widths[i]
    ws.freeze_panes = "A2"

    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    today = datetime.date.today().isoformat()
    return StreamingResponse(buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="Health_Records_{today}.xlsx"'})


# ─── Download: Vendor Payments ────────────────────────────────────────────────
@app.get("/api/download/vendors")
def download_vendors(db: Session = Depends(get_db)):
    import io, openpyxl
    from openpyxl.styles import PatternFill

    vendors = db.query(models.Vendor).order_by(models.Vendor.transaction_date.desc()).all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Vendor Payments"

    headers = ["Date", "Vendor Name", "Type", "Phone",
               "Qty (L)", "Price/Unit (Rs.)", "Total (Rs.)", "Payment Status", "Notes"]
    ws.append(headers)
    hdr_fill, hdr_font, hdr_align, border = _excel_header_style()
    for cell in ws[1]:
        cell.fill = hdr_fill; cell.font = hdr_font
        cell.alignment = hdr_align; cell.border = border
    ws.row_dimensions[1].height = 22

    paid_fill   = PatternFill("solid", fgColor="D1FAE5")   # green
    unpaid_fill = PatternFill("solid", fgColor="FEE2E2")   # red

    for v in vendors:
        row = [
            str(v.transaction_date) if v.transaction_date else "",
            v.name, v.vendor_type, v.phone or "",
            round(v.quantity_liters or 0, 2),
            round(v.price_per_unit or 0, 2),
            round(v.total_amount or 0, 2),
            v.payment_status,
            v.notes or ""
        ]
        ws.append(row)
        fill = paid_fill if v.payment_status == "PAID" else unpaid_fill
        _apply_row_style(ws, fill, border)

    col_widths = [12, 24, 16, 14, 10, 16, 14, 16, 30]
    for i, col in enumerate(ws.columns):
        ws.column_dimensions[col[0].column_letter].width = col_widths[i]
    ws.freeze_panes = "A2"

    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    today = datetime.date.today().isoformat()
    return StreamingResponse(buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="Vendor_Payments_{today}.xlsx"'})


# ─── Login Log: View & Download ───────────────────────────────────────────────
@app.get("/api/login-logs", response_model=List[schemas.LoginLogSchema])
def get_login_logs(db: Session = Depends(get_db)):
    return db.query(models.LoginLog).order_by(models.LoginLog.created_at.desc()).all()


@app.get("/api/download/login-logs")
def download_login_logs(db: Session = Depends(get_db)):
    import io, openpyxl
    from openpyxl.styles import PatternFill

    logs = db.query(models.LoginLog).order_by(models.LoginLog.created_at.desc()).all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Login Audit Log"

    headers = ["#", "Partner Name", "Login Date", "Login Time", "Day of Week"]
    ws.append(headers)
    hdr_fill, hdr_font, hdr_align, border = _excel_header_style()
    for cell in ws[1]:
        cell.fill = hdr_fill; cell.font = hdr_font
        cell.alignment = hdr_align; cell.border = border
    ws.row_dimensions[1].height = 22

    alt1 = PatternFill("solid", fgColor="F0F9FF")
    alt2 = PatternFill("solid", fgColor="FFFFFF")
    for i, log in enumerate(logs):
        day_name = ""
        if log.login_date:
            day_name = log.login_date.strftime("%A")
        row = [
            i + 1,
            log.partner_name,
            str(log.login_date) if log.login_date else "",
            log.login_time or "",
            day_name
        ]
        ws.append(row)
        _apply_row_style(ws, alt1 if i % 2 == 0 else alt2, border)

    col_widths = [6, 24, 14, 12, 14]
    for i, col in enumerate(ws.columns):
        ws.column_dimensions[col[0].column_letter].width = col_widths[i]
    ws.freeze_panes = "A2"

    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    today = datetime.date.today().isoformat()
    return StreamingResponse(buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="Login_Audit_{today}.xlsx"'})


# ─── Vendor Balance Ledger ────────────────────────────────────────────────────

@app.post("/api/vendor-payments", response_model=schemas.VendorPaymentSchema)
def add_vendor_payment(payment: schemas.VendorPaymentCreate, db: Session = Depends(get_db)):
    """Record a payment installment against a vendor's outstanding balance."""
    pay_date = payment.payment_date or datetime.date.today()
    db_pay = models.VendorPayment(
        vendor_id=payment.vendor_id,
        payment_date=pay_date,
        amount_paid=payment.amount_paid,
        notes=payment.notes
    )
    db.add(db_pay)

    # Auto-update vendor payment_status
    vendor = db.query(models.Vendor).filter(models.Vendor.id == payment.vendor_id).first()
    if vendor:
        already_paid = sum(p.amount_paid for p in vendor.payments)
        new_total = already_paid + payment.amount_paid
        if new_total >= (vendor.total_amount or 0):
            vendor.payment_status = "PAID"

    db.commit()
    db.refresh(db_pay)
    return db_pay


@app.get("/api/vendor-payments/{vendor_id}", response_model=List[schemas.VendorPaymentSchema])
def get_vendor_payments(vendor_id: str, db: Session = Depends(get_db)):
    return (db.query(models.VendorPayment)
              .filter(models.VendorPayment.vendor_id == vendor_id)
              .order_by(models.VendorPayment.payment_date.desc())
              .all())


@app.get("/api/vendor-balances")
def get_vendor_balances(db: Session = Depends(get_db)):
    """Return balance summary for every vendor (total owed, paid so far, outstanding)."""
    vendors = db.query(models.Vendor).order_by(models.Vendor.transaction_date.desc()).all()
    result = []
    for v in vendors:
        total_amount = v.total_amount or 0.0
        total_paid   = sum(p.amount_paid for p in v.payments)
        outstanding  = max(0.0, total_amount - total_paid)
        result.append({
            "vendor_id":        v.id,
            "vendor_name":      v.name,
            "vendor_type":      v.vendor_type,
            "phone":            v.phone or "",
            "transaction_date": str(v.transaction_date) if v.transaction_date else "",
            "total_amount":     round(total_amount, 2),
            "total_paid":       round(total_paid, 2),
            "outstanding":      round(outstanding, 2),
            "payment_status":   "PAID" if outstanding <= 0 else "UNPAID",
        })
    return result


@app.delete("/api/vendor-payments/{payment_id}")
def delete_vendor_payment(payment_id: str, db: Session = Depends(get_db)):
    p = db.query(models.VendorPayment).filter(models.VendorPayment.id == payment_id).first()
    if not p:
        raise HTTPException(404, "Payment not found")
    vendor = db.query(models.Vendor).filter(models.Vendor.id == p.vendor_id).first()
    db.delete(p)
    db.commit()
    # Re-evaluate payment status
    if vendor:
        paid = sum(px.amount_paid for px in vendor.payments)
        vendor.payment_status = "PAID" if paid >= (vendor.total_amount or 0) else "UNPAID"
        db.commit()
    return {"message": "Payment deleted"}


# ─── Monthly Summary Report (Excel) ──────────────────────────────────────────

@app.get("/api/download/monthly-report")
def download_monthly_report(month: str = None, db: Session = Depends(get_db)):
    """
    Generate a comprehensive multi-sheet monthly Excel report.
    month param format: YYYY-MM (defaults to current month)
    """
    import io, openpyxl, calendar
    from openpyxl.styles import PatternFill, Font, Alignment

    # ── Parse month ────────────────────────────────────────────────────────────
    if not month:
        today = datetime.date.today()
        month = today.strftime("%Y-%m")
    try:
        year, mon = int(month.split("-")[0]), int(month.split("-")[1])
    except Exception:
        raise HTTPException(400, "Invalid month format. Use YYYY-MM")

    first_day = datetime.date(year, mon, 1)
    last_day  = datetime.date(year, mon, calendar.monthrange(year, mon)[1])
    month_label = first_day.strftime("%B %Y")

    # ── Data queries ───────────────────────────────────────────────────────────
    milk_records = (db.query(models.MilkProduction)
                      .filter(models.MilkProduction.date >= first_day,
                              models.MilkProduction.date <= last_day)
                      .order_by(models.MilkProduction.date).all())
    sales        = (db.query(models.MilkSales)
                      .filter(models.MilkSales.date >= first_day,
                              models.MilkSales.date <= last_day)
                      .order_by(models.MilkSales.date).all())
    expenses     = (db.query(models.Expense)
                      .filter(models.Expense.date >= first_day,
                              models.Expense.date <= last_day)
                      .order_by(models.Expense.date).all())
    vendors      = db.query(models.Vendor).all()
    bufmap       = {b.id: b.tag_number for b in db.query(models.Buffalo).all()}

    total_milk   = sum(r.total_milk_liters or 0 for r in milk_records)
    total_income = sum(s.total_income or 0 for s in sales)
    total_exp    = sum(e.amount or 0 for e in expenses)
    profit       = total_income - total_exp

    # ── Workbook setup ────────────────────────────────────────────────────────
    wb = openpyxl.Workbook()
    hdr_fill, hdr_font, hdr_align, border = _excel_header_style()

    def style_header(ws, headers, col_widths):
        ws.append(headers)
        for cell in ws[1]:
            cell.fill = hdr_fill; cell.font = hdr_font
            cell.alignment = hdr_align; cell.border = border
        ws.row_dimensions[1].height = 22
        for i, col in enumerate(ws.columns):
            ws.column_dimensions[col[0].column_letter].width = col_widths[i]
        ws.freeze_panes = "A2"

    alt1 = PatternFill("solid", fgColor="F0FDF4")
    alt2 = PatternFill("solid", fgColor="FFFFFF")
    red_fill  = PatternFill("solid", fgColor="FEE2E2")
    blue_fill = PatternFill("solid", fgColor="EFF6FF")

    # ── Sheet 1: Summary Dashboard ─────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Monthly Summary"
    title_font = Font(bold=True, size=14, color="064E3B")
    ws1["A1"] = f"Monthly Farm Report — {month_label}"
    ws1["A1"].font = title_font
    ws1.merge_cells("A1:C1")
    ws1.row_dimensions[1].height = 30

    kv_pairs = [
        ("Total Buffaloes",      db.query(models.Buffalo).count()),
        ("Milk Produced (L)",    round(total_milk, 2)),
        ("Milk Sales Records",   len(sales)),
        ("Total Income (₹)",    round(total_income, 2)),
        ("Total Expenses (₹)",  round(total_exp, 2)),
        ("Net Profit / Loss (₹)", round(profit, 2)),
        ("Expense Records",      len(expenses)),
        ("Active Vendors",       len(vendors)),
    ]
    for i, (label, val) in enumerate(kv_pairs):
        row = i + 3
        ws1.cell(row=row, column=1, value=label).font = Font(bold=True)
        ws1.cell(row=row, column=2, value=val)
        fill = PatternFill("solid", fgColor="D1FAE5" if val >= 0 or not isinstance(val, float) else "FEE2E2")
        ws1.cell(row=row, column=2).fill = fill
    ws1.column_dimensions["A"].width = 26
    ws1.column_dimensions["B"].width = 18

    # ── Sheet 2: Milk Production ───────────────────────────────────────────────
    ws2 = wb.create_sheet("Milk Production")
    style_header(ws2, ["Date", "Buffalo Tag", "Morning (L)", "Evening (L)", "Total (L)"],
                 [12, 16, 14, 14, 12])
    for i, r in enumerate(milk_records):
        ws2.append([str(r.date), bufmap.get(r.buffalo_id, r.buffalo_id[:8]),
                    round(r.morning_milk_liters, 2), round(r.evening_milk_liters, 2),
                    round(r.total_milk_liters, 2)])
        _apply_row_style(ws2, alt1 if i % 2 == 0 else alt2, border)
    # Totals row
    ws2.append(["", "TOTAL", "", "", round(total_milk, 2)])
    for cell in ws2[ws2.max_row]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="D1FAE5")

    # ── Sheet 3: Income (Milk Sales) ───────────────────────────────────────────
    ws3 = wb.create_sheet("Income")
    style_header(ws3, ["Date", "Center / Buyer", "Qty (L)", "₹/Liter", "Total Income (₹)"],
                 [12, 26, 12, 12, 18])
    for i, s in enumerate(sales):
        ws3.append([str(s.date), s.milk_center_name,
                    round(s.quantity_supplied_liters, 2), round(s.price_per_liter, 2),
                    round(s.total_income, 2)])
        _apply_row_style(ws3, alt1 if i % 2 == 0 else alt2, border)
    ws3.append(["", "TOTAL", "", "", round(total_income, 2)])
    for cell in ws3[ws3.max_row]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="D1FAE5")

    # ── Sheet 4: Expenses ─────────────────────────────────────────────────────
    ws4 = wb.create_sheet("Expenses")
    style_header(ws4, ["Date", "Category", "Amount (₹)", "Description"],
                 [12, 26, 16, 36])
    for i, e in enumerate(expenses):
        ws4.append([str(e.date), e.category, round(e.amount, 2), e.description or ""])
        _apply_row_style(ws4, red_fill if i % 2 == 0 else PatternFill("solid", fgColor="FFF5F5"), border)
    ws4.append(["", "TOTAL", round(total_exp, 2), ""])
    for cell in ws4[ws4.max_row]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="FCA5A5")

    # ── Sheet 5: Vendor Payments ───────────────────────────────────────────────
    ws5 = wb.create_sheet("Vendor Payments")
    style_header(ws5, ["Vendor", "Type", "Transaction Date", "Total Billed (₹)",
                        "Total Paid (₹)", "Outstanding (₹)", "Status"],
                 [22, 16, 16, 16, 16, 16, 12])
    total_billed = total_outstanding = 0.0
    for i, v in enumerate(vendors):
        paid = sum(p.amount_paid for p in v.payments)
        owed = max(0.0, (v.total_amount or 0) - paid)
        total_billed += v.total_amount or 0
        total_outstanding += owed
        ws5.append([v.name, v.vendor_type, str(v.transaction_date) if v.transaction_date else "",
                    round(v.total_amount or 0, 2), round(paid, 2), round(owed, 2),
                    "PAID" if owed <= 0 else "UNPAID"])
        fill = (PatternFill("solid", fgColor="D1FAE5") if owed <= 0
                else PatternFill("solid", fgColor="FEE2E2"))
        _apply_row_style(ws5, fill, border)
    ws5.append(["TOTAL", "", "", round(total_billed, 2), "", round(total_outstanding, 2), ""])
    for cell in ws5[ws5.max_row]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="FCD34D")

    # ── Stream ────────────────────────────────────────────────────────────────
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    safe = month.replace("-", "_")
    return StreamingResponse(buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="Monthly_Report_{safe}.xlsx"'})


@app.get("/api/monthly-summary")
def get_monthly_summary(month: str = None, db: Session = Depends(get_db)):
    """Return JSON summary stats for a given YYYY-MM month (default = current)."""
    import calendar
    if not month:
        today = datetime.date.today()
        month = today.strftime("%Y-%m")
    year, mon = int(month.split("-")[0]), int(month.split("-")[1])
    first_day = datetime.date(year, mon, 1)
    last_day  = datetime.date(year, mon, calendar.monthrange(year, mon)[1])

    milk_total = sum(
        r[0] for r in db.execute(
            __import__("sqlalchemy").text(
                "SELECT total_milk_liters FROM milk_production WHERE date>=:f AND date<=:l"),
            {"f": str(first_day), "l": str(last_day)}).fetchall() if r[0])
    income = sum(
        r[0] for r in db.execute(
            __import__("sqlalchemy").text(
                "SELECT total_income FROM milk_sales WHERE date>=:f AND date<=:l"),
            {"f": str(first_day), "l": str(last_day)}).fetchall() if r[0])
    expenses = sum(
        r[0] for r in db.execute(
            __import__("sqlalchemy").text(
                "SELECT amount FROM expenses WHERE date>=:f AND date<=:l"),
            {"f": str(first_day), "l": str(last_day)}).fetchall() if r[0])
    unpaid_vendors = db.query(models.Vendor).filter(models.Vendor.payment_status == "UNPAID").count()
    return {
        "month": month,
        "month_label": first_day.strftime("%B %Y"),
        "total_milk_liters": round(milk_total, 2),
        "total_income": round(income, 2),
        "total_expenses": round(expenses, 2),
        "net_profit": round(income - expenses, 2),
        "unpaid_vendors": unpaid_vendors,
    }
